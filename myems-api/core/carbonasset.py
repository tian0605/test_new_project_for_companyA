import cgi
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import falcon
import mysql.connector
from openpyxl import load_workbook
import simplejson as json

from core.useractivity import access_control, get_request_context_value
from core import utilities
import config


def _connect():
    return mysql.connector.connect(**config.myems_production_db)


def _connect_system():
    return mysql.connector.connect(**config.myems_system_db)


def _connect_carbon():
    return mysql.connector.connect(**config.myems_carbon_db)


def _authenticate(req):
    access_control(req)
    enterprise_space_id = get_request_context_value(req, 'enterprise_space_id')
    return int(enterprise_space_id) if enterprise_space_id is not None else 0


def _get_authorized_space_ids(req, enterprise_space_id):
    authorized_space_ids = get_request_context_value(req, 'authorized_space_ids')
    if isinstance(authorized_space_ids, list):
        scoped_space_ids = [space_id for space_id in authorized_space_ids if isinstance(space_id, int) and space_id > 0]
        if scoped_space_ids:
            return scoped_space_ids
    return [enterprise_space_id] if enterprise_space_id > 0 else []


def _get_timezone_offset():
    timezone_offset = int(config.utc_offset[1:3]) * 60 + int(config.utc_offset[4:6])
    if config.utc_offset[0] == '-':
        timezone_offset = -timezone_offset
    return timedelta(minutes=timezone_offset)


def _get_local_year_window_utc(accounting_year):
    timezone_delta = _get_timezone_offset()
    local_start_datetime = datetime(accounting_year, 1, 1, 0, 0, 0)
    local_end_datetime = datetime(accounting_year + 1, 1, 1, 0, 0, 0)
    return local_start_datetime - timezone_delta, local_end_datetime - timezone_delta


def _get_space_emissions_by_year(space_id, accounting_year):
    start_datetime_utc, end_datetime_utc = _get_local_year_window_utc(accounting_year)

    cnx = None
    cursor = None
    try:
        cnx = _connect_carbon()
        cursor = cnx.cursor()
        cursor.execute(
            ' SELECT start_datetime_utc, actual_value '
            ' FROM tbl_space_input_category_hourly '
            ' WHERE space_id = %s AND start_datetime_utc >= %s AND start_datetime_utc < %s '
            ' ORDER BY start_datetime_utc ',
            (space_id, start_datetime_utc, end_datetime_utc)
        )
        rows_hourly = cursor.fetchall()
    finally:
        if cursor:
            cursor.close()
        if cnx:
            cnx.close()

    annual_emissions_kg = sum((row[1] for row in rows_hourly if row[1] is not None), Decimal('0'))
    rows_monthly = utilities.aggregate_hourly_data_by_period(rows_hourly,
                                                             start_datetime_utc,
                                                             end_datetime_utc,
                                                             'monthly')
    monthly_emissions_kg = [Decimal('0') if row[1] is None else row[1] for row in rows_monthly[:12]]
    if len(monthly_emissions_kg) < 12:
        monthly_emissions_kg.extend([Decimal('0')] * (12 - len(monthly_emissions_kg)))

    annual_emissions = annual_emissions_kg / Decimal('1000')
    monthly_emissions = [value / Decimal('1000') for value in monthly_emissions_kg]

    return annual_emissions, monthly_emissions


def _read_json(req):
    try:
        raw_json = req.stream.read().decode('utf-8')
    except UnicodeDecodeError:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_ENCODING')
    except Exception:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.FAILED_TO_READ_REQUEST_STREAM')

    try:
        parsed = json.loads(raw_json)
    except Exception:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_JSON')

    if not isinstance(parsed, dict) or not isinstance(parsed.get('data'), dict):
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')
    return parsed['data']


def _read_uploaded_file(req, field_name='file'):
    try:
        raw_content_type = req.get_header('Content-Type') or req.get_header('content-type') or ''
        environ = dict(req.env)
        environ['REQUEST_METHOD'] = req.method
        environ['CONTENT_TYPE'] = raw_content_type
        environ['CONTENT_LENGTH'] = str(req.content_length or '0')
        form = cgi.FieldStorage(fp=req.env.get('wsgi.input', req.bounded_stream), environ=environ, keep_blank_values=True)
        field = form[field_name] if field_name in form else None
        if isinstance(field, list):
            field = field[0]
        if field is not None and getattr(field, 'file', None) is not None:
            return field.filename or '', field.file.read()
    except Exception as ex:
        print('carbon market upload cgi parse failed:', str(ex))

    try:
        media = req.get_media(default_when_empty=None)
    except Exception:
        media = None

    if media is not None:
        try:
            for part in media:
                if getattr(part, 'name', None) != field_name:
                    continue
                raw_blob = part.data
                filename = getattr(part, 'filename', '') or ''
                if raw_blob is None:
                    break
                return filename, raw_blob
        except Exception as ex:
            print('carbon market upload media parse failed:', str(ex))

    try:
        upload = req.get_param(field_name)
        if upload is not None:
            return upload.filename, upload.file.read()
    except Exception as ex:
        print('carbon market upload legacy param failed:', str(ex))

    raise falcon.HTTPError(status=falcon.HTTP_400,
                           title='API.ERROR',
                           description='API.FAILED_TO_UPLOAD_ATTACHMENT_FILE')


def _string_value(data, key, error, required=True, max_length=None):
    value = data.get(key)
    if value is None:
        if required:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
        return ''
    value = str(value).strip()
    if required and len(value) == 0:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    if max_length is not None and len(value) > max_length:
        value = value[:max_length]
    return value


def _decimal_value(data, key, error, required=True, minimum=None):
    value = data.get(key)
    if value is None or value == '':
        if required:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
        return Decimal('0')
    try:
        result = Decimal(str(value).replace(',', '').replace('%', ''))
    except (InvalidOperation, ValueError, TypeError):
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    if minimum is not None and result < Decimal(str(minimum)):
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    return result


def _int_value(data, key, error, required=True, minimum=None):
    value = data.get(key)
    if value is None or value == '':
        if required:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
        return None
    try:
        result = int(value)
    except (ValueError, TypeError):
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    if minimum is not None and result < minimum:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    return result


def _date_value(value, error):
    if value is None or value == '':
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
    except ValueError:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description=error)


def _normalize_market_header(value):
    if value is None:
        return ''
    normalized = str(value).strip()
    translation_table = str.maketrans({
        '（': '(',
        '）': ')',
        '％': '%',
        '，': ',',
        '　': '',
        ' ': ''
    })
    return normalized.translate(translation_table)


MARKET_HEADER_ALIASES = {
    '日期': 'trade_date',
    '交易日期': 'trade_date',
    '品种': 'variety_code',
    '品种代码': 'variety_code',
    '开盘价(元)': 'open_price',
    '开盘价': 'open_price',
    '收盘价(元)': 'close_price',
    '收盘价': 'close_price',
    '最高价(元)': 'high_price',
    '最高价': 'high_price',
    '最低价(元)': 'low_price',
    '最低价': 'low_price',
    '涨跌(元)': 'change_value',
    '涨跌': 'change_value',
    '涨跌幅(%)': 'change_rate',
    '涨跌幅': 'change_rate',
    '成交数量(吨)': 'trading_volume',
    '成交数量': 'trading_volume',
    '成交金额(元)': 'trading_amount',
    '成交金额': 'trading_amount'
}


REQUIRED_MARKET_FIELDS = [
    'trade_date', 'variety_code', 'open_price', 'close_price', 'high_price',
    'low_price', 'change_value', 'change_rate', 'trading_volume', 'trading_amount'
]


def _get_space_names(space_ids):
    if not space_ids:
        return {}
    cnx = None
    cursor = None
    try:
        cnx = _connect_system()
        cursor = cnx.cursor()
        placeholders = ','.join(['%s'] * len(space_ids))
        cursor.execute(
            f' SELECT id, name FROM tbl_spaces WHERE id IN ({placeholders}) ',
            tuple(space_ids)
        )
        return {row[0]: row[1] for row in cursor.fetchall()}
    finally:
        if cursor:
            cursor.close()
        if cnx:
            cnx.close()


def _ensure_space_authorized(space_id, authorized_space_ids):
    if space_id not in authorized_space_ids:
        raise falcon.HTTPError(status=falcon.HTTP_404, title='API.NOT_FOUND', description='API.SPACE_NOT_FOUND')


def _compute_totals(data):
    quota_total = data['government_quota'] + data['previous_year_quota'] + data['purchased_quota'] - data['sold_quota']
    ccer_total = data['own_ccer'] + data['purchased_ccer'] - data['sold_ccer']
    green_total = data['purchased_green_certificate'] - data['sold_green_certificate'] - data['retired_green_certificate']
    return quota_total, ccer_total, green_total


def _row_to_asset(row, space_name=''):
    return {
        'id': row[0],
        'uuid': row[1],
        'space_id': row[2],
        'space_name': space_name,
        'accounting_year': row[3],
        'government_quota': row[4],
        'previous_year_quota': row[5],
        'purchased_quota': row[6],
        'sold_quota': row[7],
        'own_ccer': row[8],
        'purchased_ccer': row[9],
        'sold_ccer': row[10],
        'purchased_green_certificate': row[11],
        'sold_green_certificate': row[12],
        'retired_green_certificate': row[13],
        'quota_total': row[14],
        'ccer_total': row[15],
        'green_certificate_total': row[16],
        'data_status': row[17],
        'remark': row[18]
    }


def _row_to_market_history(row):
    return {
        'id': row[0],
        'trade_date': row[1].isoformat() if row[1] else None,
        'market_code': row[2],
        'variety_code': row[3],
        'open_price': row[4],
        'close_price': row[5],
        'high_price': row[6],
        'low_price': row[7],
        'change_value': row[8],
        'change_rate': row[9],
        'trading_volume': row[10],
        'trading_amount': row[11],
        'source_name': row[12],
        'source_file_name': row[13],
        'import_batch_id': row[14],
        'import_datetime_utc': row[15].isoformat() if row[15] else None
    }


def _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids):
    cursor.execute(
        ' SELECT id, space_id FROM tbl_carbon_assets WHERE id = %s AND enterprise_space_id = %s ',
        (asset_id, enterprise_space_id)
    )
    row = cursor.fetchone()
    if row is None:
        raise falcon.HTTPError(status=falcon.HTTP_404, title='API.NOT_FOUND', description='API.CARBON_ASSET_NOT_FOUND')
    _ensure_space_authorized(row[1], authorized_space_ids)
    return row[1]


def _validate_unique_year(cursor, enterprise_space_id, space_id, accounting_year, current_id=None):
    params = [enterprise_space_id, space_id, accounting_year]
    query = ' SELECT id FROM tbl_carbon_assets WHERE enterprise_space_id = %s AND space_id = %s AND accounting_year = %s '
    if current_id is not None:
        query += ' AND id != %s '
        params.append(current_id)
    cursor.execute(query, tuple(params))
    if cursor.fetchone() is not None:
        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.CARBON_ASSET_YEAR_EXISTS')


class CarbonAssetCollection:
    @staticmethod
    def on_options(req, resp):
        _ = req
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_get(req, resp):
        enterprise_space_id = _authenticate(req)
        accounting_year = req.get_param_as_int('year', required=False)
        space_id = req.get_param_as_int('spaceid', required=False)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        if not authorized_space_ids:
            resp.text = json.dumps([], use_decimal=True)
            return
        if space_id is not None:
            _ensure_space_authorized(space_id, authorized_space_ids)

        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            placeholders = ','.join(['%s'] * len(authorized_space_ids))
            query = (' SELECT id, uuid, space_id, accounting_year, government_quota, previous_year_quota, purchased_quota, sold_quota, '
                     '        own_ccer, purchased_ccer, sold_ccer, purchased_green_certificate, sold_green_certificate, '
                     '        retired_green_certificate, quota_total, ccer_total, green_certificate_total, data_status, remark '
                     ' FROM tbl_carbon_assets WHERE enterprise_space_id = %s '
                     f' AND space_id IN ({placeholders}) ')
            params = [enterprise_space_id] + authorized_space_ids
            if accounting_year is not None:
                query += ' AND accounting_year = %s '
                params.append(accounting_year)
            if space_id is not None:
                query += ' AND space_id = %s '
                params.append(space_id)
            query += ' ORDER BY accounting_year DESC, space_id '
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()
            space_names = _get_space_names([row[2] for row in rows])
            result = [_row_to_asset(row, space_names.get(row[2], '')) for row in rows]
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        resp.text = json.dumps(result, use_decimal=True)

    @staticmethod
    def on_post(req, resp):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        data = _read_json(req)
        space_id = _int_value(data, 'space_id', 'API.INVALID_SPACE_ID', minimum=1)
        _ensure_space_authorized(space_id, authorized_space_ids)
        accounting_year = _int_value(data, 'accounting_year', 'API.INVALID_YEAR', minimum=1900)
        payload = {
            'government_quota': _decimal_value(data, 'government_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'previous_year_quota': _decimal_value(data, 'previous_year_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'purchased_quota': _decimal_value(data, 'purchased_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'sold_quota': _decimal_value(data, 'sold_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'own_ccer': _decimal_value(data, 'own_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'purchased_ccer': _decimal_value(data, 'purchased_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'sold_ccer': _decimal_value(data, 'sold_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'purchased_green_certificate': _decimal_value(data, 'purchased_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0'),
            'sold_green_certificate': _decimal_value(data, 'sold_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0'),
            'retired_green_certificate': _decimal_value(data, 'retired_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0')
        }
        quota_total, ccer_total, green_total = _compute_totals(payload)
        data_status = _string_value(data, 'data_status', 'API.INVALID_DATA_STATUS', required=False, max_length=32) or 'active'
        remark = _string_value(data, 'remark', 'API.INVALID_REMARK', required=False, max_length=255)

        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _validate_unique_year(cursor, enterprise_space_id, space_id, accounting_year)
            cursor.execute(
                ' INSERT INTO tbl_carbon_assets '
                ' (uuid, enterprise_space_id, space_id, accounting_year, government_quota, previous_year_quota, purchased_quota, sold_quota, '
                '   own_ccer, purchased_ccer, sold_ccer, purchased_green_certificate, sold_green_certificate, retired_green_certificate, '
                '   quota_total, ccer_total, green_certificate_total, data_status, remark) '
                ' VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ',
                (str(uuid.uuid4()), enterprise_space_id, space_id, accounting_year, payload['government_quota'], payload['previous_year_quota'],
                 payload['purchased_quota'], payload['sold_quota'], payload['own_ccer'], payload['purchased_ccer'], payload['sold_ccer'],
                 payload['purchased_green_certificate'], payload['sold_green_certificate'], payload['retired_green_certificate'],
                 quota_total, ccer_total, green_total, data_status, remark)
            )
            new_id = cursor.lastrowid
            cnx.commit()
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        resp.status = falcon.HTTP_201
        resp.location = '/carbon-assets/' + str(new_id)


class CarbonAssetItem:
    @staticmethod
    def on_options(req, resp, id_):
        _ = req
        _ = id_
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_get(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            cursor.execute(
                ' SELECT id, uuid, space_id, accounting_year, government_quota, previous_year_quota, purchased_quota, sold_quota, '
                '        own_ccer, purchased_ccer, sold_ccer, purchased_green_certificate, sold_green_certificate, '
                '        retired_green_certificate, quota_total, ccer_total, green_certificate_total, data_status, remark '
                ' FROM tbl_carbon_assets WHERE id = %s AND enterprise_space_id = %s ',
                (asset_id, enterprise_space_id)
            )
            row = cursor.fetchone()
            result = _row_to_asset(row, _get_space_names([row[2]]).get(row[2], ''))
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.text = json.dumps(result, use_decimal=True)

    @staticmethod
    def on_put(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        data = _read_json(req)
        space_id = _int_value(data, 'space_id', 'API.INVALID_SPACE_ID', minimum=1)
        _ensure_space_authorized(space_id, authorized_space_ids)
        accounting_year = _int_value(data, 'accounting_year', 'API.INVALID_YEAR', minimum=1900)
        payload = {
            'government_quota': _decimal_value(data, 'government_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'previous_year_quota': _decimal_value(data, 'previous_year_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'purchased_quota': _decimal_value(data, 'purchased_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'sold_quota': _decimal_value(data, 'sold_quota', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0'),
            'own_ccer': _decimal_value(data, 'own_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'purchased_ccer': _decimal_value(data, 'purchased_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'sold_ccer': _decimal_value(data, 'sold_ccer', 'API.INVALID_CARBON_ASSET_CCER', required=False, minimum='0'),
            'purchased_green_certificate': _decimal_value(data, 'purchased_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0'),
            'sold_green_certificate': _decimal_value(data, 'sold_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0'),
            'retired_green_certificate': _decimal_value(data, 'retired_green_certificate', 'API.INVALID_CARBON_ASSET_GREEN_CERTIFICATE', required=False, minimum='0')
        }
        quota_total, ccer_total, green_total = _compute_totals(payload)
        data_status = _string_value(data, 'data_status', 'API.INVALID_DATA_STATUS', required=False, max_length=32) or 'active'
        remark = _string_value(data, 'remark', 'API.INVALID_REMARK', required=False, max_length=255)

        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            _validate_unique_year(cursor, enterprise_space_id, space_id, accounting_year, asset_id)
            cursor.execute(
                ' UPDATE tbl_carbon_assets SET '
                ' space_id = %s, accounting_year = %s, government_quota = %s, previous_year_quota = %s, purchased_quota = %s, sold_quota = %s, '
                ' own_ccer = %s, purchased_ccer = %s, sold_ccer = %s, purchased_green_certificate = %s, sold_green_certificate = %s, '
                ' retired_green_certificate = %s, quota_total = %s, ccer_total = %s, green_certificate_total = %s, data_status = %s, remark = %s '
                ' WHERE id = %s AND enterprise_space_id = %s ',
                (space_id, accounting_year, payload['government_quota'], payload['previous_year_quota'], payload['purchased_quota'],
                 payload['sold_quota'], payload['own_ccer'], payload['purchased_ccer'], payload['sold_ccer'], payload['purchased_green_certificate'],
                 payload['sold_green_certificate'], payload['retired_green_certificate'], quota_total, ccer_total, green_total, data_status, remark,
                 asset_id, enterprise_space_id)
            )
            cnx.commit()
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_delete(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            cursor.execute(' DELETE FROM tbl_carbon_asset_monthly_quotas WHERE carbon_asset_id = %s ', (asset_id,))
            cursor.execute(' DELETE FROM tbl_carbon_assets WHERE id = %s AND enterprise_space_id = %s ', (asset_id, enterprise_space_id))
            cnx.commit()
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.status = falcon.HTTP_204


class CarbonAssetMonthlyQuotaCollection:
    @staticmethod
    def on_options(req, resp, id_):
        _ = req
        _ = id_
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_get(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            cursor.execute(
                ' SELECT month_of_year, quota_amount, remark FROM tbl_carbon_asset_monthly_quotas WHERE carbon_asset_id = %s ORDER BY month_of_year ',
                (asset_id,)
            )
            rows = cursor.fetchall()
            monthly_map = {row[0]: {'month_of_year': row[0], 'quota_amount': row[1], 'remark': row[2]} for row in rows}
            result = [monthly_map.get(month, {'month_of_year': month, 'quota_amount': Decimal('0'), 'remark': ''}) for month in range(1, 13)]
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.text = json.dumps(result, use_decimal=True)

    @staticmethod
    def on_put(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        data = _read_json(req)
        monthly_quotas = data.get('monthly_quotas')
        if not isinstance(monthly_quotas, list):
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')

        normalized_rows = []
        total = Decimal('0')
        seen_months = set()
        for item in monthly_quotas:
            if not isinstance(item, dict):
                raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')
            month_of_year = _int_value(item, 'month_of_year', 'API.INVALID_MONTH', minimum=1)
            if month_of_year > 12 or month_of_year in seen_months:
                raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_MONTH')
            seen_months.add(month_of_year)
            quota_amount = _decimal_value(item, 'quota_amount', 'API.INVALID_CARBON_ASSET_QUOTA', required=False, minimum='0')
            remark = _string_value(item, 'remark', 'API.INVALID_REMARK', required=False, max_length=255)
            total += quota_amount
            normalized_rows.append((month_of_year, quota_amount, remark))

        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            cursor.execute(' SELECT quota_total FROM tbl_carbon_assets WHERE id = %s AND enterprise_space_id = %s ',
                           (asset_id, enterprise_space_id))
            asset_quota_total = cursor.fetchone()[0]
            if total != asset_quota_total:
                raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.CARBON_ASSET_MONTHLY_TOTAL_MISMATCH')
            cursor.execute(' DELETE FROM tbl_carbon_asset_monthly_quotas WHERE carbon_asset_id = %s ', (asset_id,))
            for month_of_year, quota_amount, remark in normalized_rows:
                cursor.execute(
                    ' INSERT INTO tbl_carbon_asset_monthly_quotas (carbon_asset_id, month_of_year, quota_amount, remark) VALUES (%s, %s, %s, %s) ',
                    (asset_id, month_of_year, quota_amount, remark)
                )
            cnx.commit()
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.status = falcon.HTTP_200


class CarbonAssetOverview:
    @staticmethod
    def on_options(req, resp, id_):
        _ = req
        _ = id_
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_get(req, resp, id_):
        enterprise_space_id = _authenticate(req)
        authorized_space_ids = _get_authorized_space_ids(req, enterprise_space_id)
        asset_id = _int_value({'id': id_}, 'id', 'API.INVALID_CARBON_ASSET_ID', minimum=1)
        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            _ensure_asset_exists(cursor, enterprise_space_id, asset_id, authorized_space_ids)
            cursor.execute(
                ' SELECT id, uuid, space_id, accounting_year, government_quota, previous_year_quota, purchased_quota, sold_quota, '
                '        own_ccer, purchased_ccer, sold_ccer, purchased_green_certificate, sold_green_certificate, '
                '        retired_green_certificate, quota_total, ccer_total, green_certificate_total, data_status, remark '
                ' FROM tbl_carbon_assets WHERE id = %s AND enterprise_space_id = %s ',
                (asset_id, enterprise_space_id)
            )
            row = cursor.fetchone()
            asset = _row_to_asset(row, _get_space_names([row[2]]).get(row[2], ''))
            cursor.execute(
                ' SELECT month_of_year, quota_amount FROM tbl_carbon_asset_monthly_quotas WHERE carbon_asset_id = %s ORDER BY month_of_year ',
                (asset_id,)
            )
            monthly_quota_rows = cursor.fetchall()
            monthly_quota_map = {row_item[0]: row_item[1] for row_item in monthly_quota_rows}
            monthly_quotas = [monthly_quota_map.get(month, Decimal('0')) for month in range(1, 13)]
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        annual_emissions, monthly_emissions = _get_space_emissions_by_year(asset['space_id'], asset['accounting_year'])
        result = {
            'asset': asset,
            'cards': {
                'quota_total': asset['quota_total'],
                'ccer_total': asset['ccer_total'],
                'green_certificate_total': asset['green_certificate_total'],
                'annual_emissions': annual_emissions,
                'remaining_allowance': asset['quota_total'] + asset['ccer_total'] - annual_emissions
            },
            'horizontal_bar': [
                {'name': 'Carbon Asset Total', 'value': asset['quota_total'] + asset['ccer_total']},
                {'name': 'Allowance Quota', 'value': asset['quota_total']},
                {'name': 'CCER', 'value': asset['ccer_total']},
                {'name': 'Annual Emissions', 'value': annual_emissions}
            ],
            'monthly_chart': {
                'labels': [f'{month}月' for month in range(1, 13)],
                'quota': monthly_quotas,
                'emissions': monthly_emissions
            }
        }
        resp.text = json.dumps(result, use_decimal=True)


class CarbonMarketHistoryCollection:
    @staticmethod
    def on_options(req, resp):
        _ = req
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_get(req, resp):
        _authenticate(req)
        start_date = req.get_param('startdate', default='').strip()
        end_date = req.get_param('enddate', default='').strip()
        variety_code = req.get_param('variety', default='').strip()
        page = req.get_param_as_int('page', required=False) or 1
        page_size = req.get_param_as_int('pagesize', required=False) or 10
        page = max(page, 1)
        page_size = min(max(page_size, 1), 200)
        offset = (page - 1) * page_size

        cnx = None
        cursor = None
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            query = ' FROM tbl_carbon_market_histories WHERE 1 = 1 '
            params = []
            if start_date:
                query += ' AND trade_date >= %s '
                params.append(_date_value(start_date, 'API.INVALID_DATE'))
            if end_date:
                query += ' AND trade_date <= %s '
                params.append(_date_value(end_date, 'API.INVALID_DATE'))
            if variety_code:
                query += ' AND variety_code = %s '
                params.append(variety_code)
            cursor.execute(' SELECT COUNT(*) ' + query, tuple(params))
            total = cursor.fetchone()[0]
            cursor.execute(
                ' SELECT id, trade_date, market_code, variety_code, open_price, close_price, high_price, low_price, change_value, '
                '        change_rate, trading_volume, trading_amount, source_name, source_file_name, import_batch_id, import_datetime_utc '
                + query + ' ORDER BY trade_date DESC, variety_code LIMIT %s OFFSET %s ',
                tuple(params + [page_size, offset])
            )
            result = [_row_to_market_history(row) for row in cursor.fetchall()]
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()
        resp.text = json.dumps({'total': total, 'page': page, 'pagesize': page_size, 'items': result}, use_decimal=True)


class CarbonMarketHistoryImport:
    @staticmethod
    def on_options(req, resp):
        _ = req
        resp.status = falcon.HTTP_200

    @staticmethod
    def on_post(req, resp):
        _authenticate(req)
        filename, raw_blob = _read_uploaded_file(req)

        try:
            workbook = load_workbook(filename=BytesIO(raw_blob), read_only=False, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
        except Exception:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')
        header_map = {}
        for index, header in enumerate(rows[0]):
            normalized_header = _normalize_market_header(header)
            field_name = MARKET_HEADER_ALIASES.get(normalized_header)
            if field_name:
                header_map[field_name] = index
        missing_fields = [field for field in REQUIRED_MARKET_FIELDS if field not in header_map]
        if missing_fields:
            raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')

        cnx = None
        cursor = None
        import_batch_id = str(uuid.uuid4())
        inserted = 0
        updated = 0
        failed = []
        try:
            cnx = _connect()
            cursor = cnx.cursor()
            for row_index, row in enumerate(rows[1:], start=2):
                if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                try:
                    trade_date = _date_value(row[header_map['trade_date']], 'API.INVALID_DATE')
                    variety_code = str(row[header_map['variety_code']]).strip()
                    if len(variety_code) == 0:
                        raise falcon.HTTPError(status=falcon.HTTP_400, title='API.BAD_REQUEST', description='API.INVALID_DATA')
                    open_price = _decimal_value({'value': row[header_map['open_price']]}, 'value', 'API.INVALID_DATA', minimum='0')
                    close_price = _decimal_value({'value': row[header_map['close_price']]}, 'value', 'API.INVALID_DATA', minimum='0')
                    raw_high_price = row[header_map['high_price']]
                    raw_low_price = row[header_map['low_price']]
                    high_price = _decimal_value({'value': raw_high_price}, 'value', 'API.INVALID_DATA', required=False, minimum='0')
                    low_price = _decimal_value({'value': raw_low_price}, 'value', 'API.INVALID_DATA', required=False, minimum='0')
                    if raw_high_price is None or str(raw_high_price).strip() == '':
                        high_price = max(open_price, close_price)
                    if raw_low_price is None or str(raw_low_price).strip() == '':
                        low_price = min(open_price, close_price)
                    change_value = _decimal_value({'value': row[header_map['change_value']]}, 'value', 'API.INVALID_DATA', required=False)
                    change_rate = _decimal_value({'value': row[header_map['change_rate']]}, 'value', 'API.INVALID_DATA', required=False)
                    trading_volume = _decimal_value({'value': row[header_map['trading_volume']]}, 'value', 'API.INVALID_DATA', required=False, minimum='0')
                    trading_amount = _decimal_value({'value': row[header_map['trading_amount']]}, 'value', 'API.INVALID_DATA', required=False, minimum='0')
                    cursor.execute(
                        ' INSERT INTO tbl_carbon_market_histories '
                        ' (uuid, trade_date, market_code, variety_code, open_price, close_price, high_price, low_price, change_value, '
                        '   change_rate, trading_volume, trading_amount, source_name, source_file_name, import_batch_id) '
                        ' VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) '
                        ' ON DUPLICATE KEY UPDATE '
                        ' open_price = VALUES(open_price), close_price = VALUES(close_price), high_price = VALUES(high_price), '
                        ' low_price = VALUES(low_price), change_value = VALUES(change_value), change_rate = VALUES(change_rate), '
                        ' trading_volume = VALUES(trading_volume), trading_amount = VALUES(trading_amount), '
                        ' source_name = VALUES(source_name), source_file_name = VALUES(source_file_name), import_batch_id = VALUES(import_batch_id), '
                        ' import_datetime_utc = CURRENT_TIMESTAMP ',
                        (str(uuid.uuid4()), trade_date, 'GZEA', variety_code, open_price, close_price, high_price, low_price,
                         change_value, change_rate, trading_volume, trading_amount, '广州碳排放权交易中心', filename, import_batch_id)
                    )
                    if cursor.rowcount == 1:
                        inserted += 1
                    else:
                        updated += 1
                except Exception as ex:
                    failed.append({'row': row_index, 'message': str(ex)})
            cnx.commit()
        finally:
            if cursor:
                cursor.close()
            if cnx:
                cnx.close()

        resp.text = json.dumps({
            'import_batch_id': import_batch_id,
            'source_file_name': filename,
            'inserted': inserted,
            'updated': updated,
            'failed': len(failed),
            'errors': failed[:20]
        }, use_decimal=True)