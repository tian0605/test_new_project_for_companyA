import base64
import os
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.utilities import round2


EXPORT_LABELS = {
    'report_title': '节能评估报告',
    'enterprise_name': '企业名称',
    'product_name': '产品名称',
    'period_type': '时间尺度',
    'reporting_start': '报告期开始时间',
    'reporting_end': '报告期结束时间',
    'summary': '汇总',
    'item': '项目',
    'value': '数值',
    'unit': '单位',
    'evaluation': '评价结果',
    'metric': '指标',
    'actual_value': '实际值',
    'benchmark': '标杆值',
    'evaluation_text': '评价说明',
    'status': '评价状态',
    'advice': '建议',
    'trend': '趋势明细',
    'datetime': '时间',
    'production': '产量',
    'ton_of_standard_coal': '吨标准煤',
    'ton_of_carbon_dioxide_emissions': '吨二氧化碳排放',
    'per_unit_energy': '单位产品综合能耗',
    'per_unit_carbon': '单位产品二氧化碳排放',
    'energy_intensity': '单位产品综合能耗',
    'carbon_intensity': '单位产品二氧化碳排放',
}


PERIOD_TYPE_LABELS = {
    'hourly': '小时',
    'daily': '日',
    'weekly': '周',
    'monthly': '月',
    'yearly': '年',
}


METRIC_NAME_LABELS = {
    'unit_comprehensive_energy_tce_per_t': EXPORT_LABELS['per_unit_energy'],
    'unit_carbon_tco2_per_t': EXPORT_LABELS['per_unit_carbon'],
    'Per Unit Product Energy Consumption': EXPORT_LABELS['per_unit_energy'],
    'Per Unit Product Carbon Dioxide Emissions': EXPORT_LABELS['per_unit_carbon'],
}


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    if report is None:
        return None

    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    base64_encoded_data = base64.b64encode(binary_file_data)
    base64_message = base64_encoded_data.decode('utf-8')

    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))

    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):
    wb = Workbook()
    ws = wb.active
    ws.title = 'SpaceEvaluation'

    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[column].width = 22
    ws.column_dimensions['A'].width = 2

    title_font = Font(name='Arial', size=15, bold=True)
    name_font = Font(name='Arial', size=12, bold=True)
    cell_font = Font(name='Arial', size=11)
    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    border = Border(left=Side(border_style='medium'),
                    right=Side(border_style='medium'),
                    top=Side(border_style='medium'),
                    bottom=Side(border_style='medium'))
    underline_border = Border(bottom=Side(border_style='medium'))
    center_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    right_alignment = Alignment(vertical='bottom', horizontal='right', wrap_text=True)

    ws['B2'].font = title_font
    ws['B2'] = name + ' ' + EXPORT_LABELS['report_title']

    ws['B4'].alignment = right_alignment
    ws['B4'] = EXPORT_LABELS['enterprise_name'] + ':'
    ws['C4'].border = underline_border
    ws['C4'].alignment = center_alignment
    ws['C4'] = name

    ws['D4'].alignment = right_alignment
    ws['D4'] = EXPORT_LABELS['product_name'] + ':'
    ws['E4'].border = underline_border
    ws['E4'].alignment = center_alignment
    ws['E4'] = report.get('product', {}).get('name') or ''

    ws['B5'].alignment = right_alignment
    ws['B5'] = EXPORT_LABELS['period_type'] + ':'
    ws['C5'].border = underline_border
    ws['C5'].alignment = center_alignment
    ws['C5'] = PERIOD_TYPE_LABELS.get(period_type, period_type)

    ws['D5'].alignment = right_alignment
    ws['D5'] = EXPORT_LABELS['reporting_start'] + ':'
    ws['E5'].border = underline_border
    ws['E5'].alignment = center_alignment
    ws['E5'] = reporting_start_datetime_local

    ws['B6'].alignment = right_alignment
    ws['B6'] = EXPORT_LABELS['reporting_end'] + ':'
    ws['C6'].border = underline_border
    ws['C6'].alignment = center_alignment
    ws['C6'] = reporting_end_datetime_local

    current_row = 8
    ws['B' + str(current_row)].font = title_font
    ws['B' + str(current_row)] = EXPORT_LABELS['summary']
    current_row += 1

    summary_headers = [EXPORT_LABELS['item'], EXPORT_LABELS['value'], EXPORT_LABELS['unit']]
    for column_index, header in enumerate(summary_headers, start=2):
        cell = ws.cell(row=current_row, column=column_index)
        cell.value = header
        cell.font = name_font
        cell.alignment = center_alignment
        cell.fill = table_fill
        cell.border = border
    current_row += 1

    summary = report.get('summary') or {}
    product = report.get('product') or {}
    summary_rows = [
        (EXPORT_LABELS['production'], summary.get('total_production'), product.get('unit') or ''),
        (EXPORT_LABELS['ton_of_standard_coal'], to_tons(summary.get('total_energy_kgce')), 'TCE'),
        (EXPORT_LABELS['ton_of_carbon_dioxide_emissions'], to_tons(summary.get('total_carbon_kgco2e')), 'TCO2E'),
        (EXPORT_LABELS['per_unit_energy'], summary.get('unit_comprehensive_energy_tce_per_t'), 'TCE/T'),
        (EXPORT_LABELS['per_unit_carbon'], summary.get('unit_carbon_tco2_per_t'), 'TCO2E/T'),
    ]
    for row_values in summary_rows:
        for column_index, value in enumerate(row_values, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = round_or_blank(value) if column_index == 3 else value
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

    current_row += 1
    ws['B' + str(current_row)].font = title_font
    ws['B' + str(current_row)] = EXPORT_LABELS['evaluation']
    current_row += 1

    evaluation_headers = [
        EXPORT_LABELS['metric'],
        EXPORT_LABELS['actual_value'],
        EXPORT_LABELS['benchmark'],
        EXPORT_LABELS['evaluation_text'],
        EXPORT_LABELS['status'],
        EXPORT_LABELS['advice']
    ]
    for column_index, header in enumerate(evaluation_headers, start=2):
        cell = ws.cell(row=current_row, column=column_index)
        cell.value = header
        cell.font = name_font
        cell.alignment = center_alignment
        cell.fill = table_fill
        cell.border = border
    current_row += 1

    for evaluation in report.get('evaluations') or []:
        values = [
            get_metric_display_name(evaluation),
            round_or_blank(evaluation.get('actual_value')),
            round_or_blank(evaluation.get('benchmark_value')),
            get_chinese_text(evaluation.get('evaluation_text'), evaluation.get('grade_label')),
            get_chinese_text(evaluation.get('status_text'), evaluation.get('grade_label')),
            evaluation.get('advice_text'),
        ]
        for column_index, value in enumerate(values, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = value
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

    current_row += 1
    ws['B' + str(current_row)].font = title_font
    ws['B' + str(current_row)] = EXPORT_LABELS['trend']
    current_row += 1

    trend_headers = [
        EXPORT_LABELS['datetime'],
        EXPORT_LABELS['production'],
        EXPORT_LABELS['energy_intensity'],
        EXPORT_LABELS['carbon_intensity']
    ]
    for column_index, header in enumerate(trend_headers, start=2):
        cell = ws.cell(row=current_row, column=column_index)
        cell.value = header
        cell.font = name_font
        cell.alignment = center_alignment
        cell.fill = table_fill
        cell.border = border
    current_row += 1

    for trend in report.get('trends') or []:
        values = [
            trend.get('datetime'),
            round_or_blank(trend.get('production')),
            round_or_blank(trend.get('unit_comprehensive_energy_tce_per_t')),
            round_or_blank(trend.get('unit_carbon_tco2_per_t')),
        ]
        for column_index, value in enumerate(values, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = value
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)
    return filename


def to_tons(value):
    return value / 1000 if value is not None else None


def round_or_blank(value):
    if value is None:
        return ''
    try:
        return round2(value)
    except Exception:
        return value


def get_metric_display_name(evaluation):
    metric_name = evaluation.get('metric_name') or evaluation.get('metric_code')
    metric_code = evaluation.get('metric_code')
    return METRIC_NAME_LABELS.get(metric_code) or METRIC_NAME_LABELS.get(metric_name) or metric_name


def get_chinese_text(value, fallback):
    if value is None:
        return fallback or ''

    text = str(value).strip()
    if not text:
        return fallback or ''

    return text
