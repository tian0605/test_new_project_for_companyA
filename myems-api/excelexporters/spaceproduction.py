import base64
import os
import uuid

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.utilities import get_translation, round2


def export(report,
           name,
           base_period_start_datetime_local,
           base_period_end_datetime_local,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    if report is None:
        return None

    filename = generate_excel(
        report,
        name,
        base_period_start_datetime_local,
        base_period_end_datetime_local,
        reporting_start_datetime_local,
        reporting_end_datetime_local,
        period_type,
        language,
    )

    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        print(str(ex))

    base64_encoded_data = base64.b64encode(binary_file_data)
    base64_message = base64_encoded_data.decode('utf-8')

    try:
        os.remove(filename)
    except NotImplementedError as ex:
        print(str(ex))

    return base64_message


def generate_excel(report,
                   name,
                   base_period_start_datetime_local,
                   base_period_end_datetime_local,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):
    trans = get_translation(language)
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = 'SpaceProduction'

    ws.row_dimensions[1].height = 102
    for i in range(2, 400 + 1):
        ws.row_dimensions[i].height = 28

    ws.column_dimensions['A'].width = 1.5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 22

    title_font = Font(name='Arial', size=15, bold=True)
    name_font = Font(name='Arial', size=12, bold=True)
    cell_font = Font(name='Arial', size=11)
    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    border = Border(
        left=Side(border_style='medium'),
        right=Side(border_style='medium'),
        top=Side(border_style='medium'),
        bottom=Side(border_style='medium'),
    )
    underline_border = Border(bottom=Side(border_style='medium'))
    center_alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    right_alignment = Alignment(vertical='bottom', horizontal='right', wrap_text=True)

    ws['B3'].alignment = right_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = underline_border
    ws['C3'].alignment = center_alignment
    ws['C3'] = name

    ws['D3'].alignment = right_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = underline_border
    ws['E3'].alignment = center_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = right_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = underline_border
    ws['C4'].alignment = center_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = right_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = underline_border
    ws['E4'].alignment = center_alignment
    ws['E4'] = reporting_end_datetime_local

    if is_base_period_timestamp_exists(report.get('base_period', {})):
        ws['B5'].alignment = right_alignment
        ws['B5'] = _('Base Period Start Datetime') + ':'
        ws['C5'].border = underline_border
        ws['C5'].alignment = center_alignment
        ws['C5'] = base_period_start_datetime_local

        ws['D5'].alignment = right_alignment
        ws['D5'] = _('Base Period End Datetime') + ':'
        ws['E5'].border = underline_border
        ws['E5'].alignment = center_alignment
        ws['E5'] = base_period_end_datetime_local

    product = report.get('product', {})
    product_name = product.get('name') or _('Production')
    product_unit = product.get('unit') or ''
    reporting_period = report.get('reporting_period', {})
    base_period = report.get('base_period', {})

    summary_rows = [
        (_('Base Period Production'), report.get('base_total_production'), product_unit, None),
        (_('Reporting Period Production'), report.get('reporting_total_production'), product_unit, None),
        (_('Ton of Standard Coal'), to_tons(reporting_period.get('total_in_kgce')), 'TCE', reporting_period.get('increment_rate_in_kgce')),
        (_('Per Unit Product Energy Consumption'), to_tons(reporting_period.get('total_in_kgce_per_prodution')), 'TCE/' + product_unit if product_unit else 'TCE', None),
        (_('Ton of Carbon Dioxide Emissions'), to_tons(reporting_period.get('total_in_kgco2e')), 'TCO2E', reporting_period.get('increment_rate_in_kgco2e')),
        (_('Per Unit Product Carbon Dioxide Emissions'), to_tons(reporting_period.get('total_in_kgco2e_per_prodution')), 'TCO2E/' + product_unit if product_unit else 'TCO2E', None),
    ]

    current_row = 7
    ws['B' + str(current_row)].font = title_font
    ws['B' + str(current_row)] = name + ' ' + _('Production')
    current_row += 1

    headers = [_('Item'), _('Value'), _('Unit'), _('Increment Rate')]
    for column_index, header in enumerate(headers, start=2):
        cell = ws.cell(row=current_row, column=column_index)
        cell.value = header
        cell.font = name_font
        cell.alignment = center_alignment
        cell.fill = table_fill
        cell.border = border
    current_row += 1

    for item_name, value, unit, increment_rate in summary_rows:
        row_values = [
            item_name,
            round_or_blank(value),
            unit,
            format_rate(increment_rate),
        ]
        for column_index, row_value in enumerate(row_values, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = row_value
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

    category_names = reporting_period.get('names') or []
    if category_names:
        current_row += 1
        ws['B' + str(current_row)].font = title_font
        ws['B' + str(current_row)] = name + ' ' + _('Production') + ' - ' + _('Energy Category')
        current_row += 1

        category_headers = [_('Energy Category'), _('Ton of Standard Coal') + ' (TCE)', _('Ton of Carbon Dioxide Emissions') + ' (TCO2E)']
        for column_index, header in enumerate(category_headers, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = header
            cell.font = name_font
            cell.alignment = center_alignment
            cell.fill = table_fill
            cell.border = border
        current_row += 1

        subtotals_in_kgce = reporting_period.get('subtotals_in_kgce') or []
        subtotals_in_kgco2e = reporting_period.get('subtotals_in_kgco2e') or []
        for index, category_name in enumerate(category_names):
            values = [
                category_name,
                round_or_blank(to_tons(subtotals_in_kgce[index]) if index < len(subtotals_in_kgce) else None),
                round_or_blank(to_tons(subtotals_in_kgco2e[index]) if index < len(subtotals_in_kgco2e) else None),
            ]
            for column_index, row_value in enumerate(values, start=2):
                cell = ws.cell(row=current_row, column=column_index)
                cell.value = row_value
                cell.font = cell_font
                cell.alignment = center_alignment
                cell.border = border
            current_row += 1

    current_row += 1
    ws['B' + str(current_row)].font = title_font
    ws['B' + str(current_row)] = name + ' ' + _('Detailed Data')
    current_row += 1

    has_base_period = is_base_period_timestamp_exists(base_period)
    if has_base_period:
        detailed_headers = [
            _('Base Period') + ' - ' + _('Datetime'),
            _('Base Period') + ' - ' + _('Production') + format_unit(product_unit),
            _('Reporting Period') + ' - ' + _('Ton of Standard Coal') + ' (TCE)',
            _('Reporting Period') + ' - ' + _('Per Unit Product Energy Consumption') + format_unit('TCE/' + product_unit if product_unit else 'TCE'),
            _('Reporting Period') + ' - ' + _('Ton of Carbon Dioxide Emissions') + ' (TCO2E)',
            _('Reporting Period') + ' - ' + _('Per Unit Product Carbon Dioxide Emissions') + format_unit('TCO2E/' + product_unit if product_unit else 'TCO2E'),
            _('Reporting Period') + ' - ' + _('Datetime'),
            _('Reporting Period') + ' - ' + _('Production') + format_unit(product_unit),
        ]
    else:
        detailed_headers = [
            _('Datetime'),
            product_name + format_unit(product_unit),
            _('Ton of Standard Coal') + ' (TCE)',
            _('Per Unit Product Energy Consumption') + format_unit('TCE/' + product_unit if product_unit else 'TCE'),
            _('Ton of Carbon Dioxide Emissions') + ' (TCO2E)',
            _('Per Unit Product Carbon Dioxide Emissions') + format_unit('TCO2E/' + product_unit if product_unit else 'TCO2E'),
        ]

    for column_index, header in enumerate(detailed_headers, start=2):
        cell = ws.cell(row=current_row, column=column_index)
        cell.value = header
        cell.font = name_font
        cell.alignment = center_alignment
        cell.fill = table_fill
        cell.border = border
    current_row += 1

    detailed_rows = build_detailed_rows(report, has_base_period)
    for row_values in detailed_rows:
        for column_index, row_value in enumerate(row_values, start=2):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = row_value
            cell.font = cell_font
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)
    return filename


def build_detailed_rows(report, has_base_period):
    reporting_production = report.get('reporting_production', {})
    reporting_timestamps = reporting_production.get('timestamps') or []
    reporting_values = reporting_production.get('values') or []
    reporting_period = report.get('reporting_period', {})
    values_in_kgce = reporting_period.get('values_in_kgce') or []
    values_in_kgce_per_production = reporting_period.get('values_in_kgce_per_production') or []
    values_in_kgco2e = reporting_period.get('values_in_kgco2e') or []
    values_in_kgco2e_per_production = reporting_period.get('values_in_kgco2e_per_production') or []

    rows = []
    if has_base_period:
        base_production = report.get('base_production', {})
        base_timestamps = base_production.get('timestamps') or []
        base_values = base_production.get('values') or []
        max_length = max(len(base_timestamps), len(reporting_timestamps))
        for index in range(max_length):
            rows.append([
                base_timestamps[index] if index < len(base_timestamps) else None,
                round_or_blank(base_values[index] if index < len(base_values) else None),
                round_or_blank(to_tons(values_in_kgce[index]) if index < len(values_in_kgce) else None),
                round_or_blank(to_tons(values_in_kgce_per_production[index]) if index < len(values_in_kgce_per_production) else None),
                round_or_blank(to_tons(values_in_kgco2e[index]) if index < len(values_in_kgco2e) else None),
                round_or_blank(to_tons(values_in_kgco2e_per_production[index]) if index < len(values_in_kgco2e_per_production) else None),
                reporting_timestamps[index] if index < len(reporting_timestamps) else None,
                round_or_blank(reporting_values[index] if index < len(reporting_values) else None),
            ])

        rows.append([
            'Subtotal',
            round_or_blank(report.get('base_total_production')),
            round_or_blank(to_tons(reporting_period.get('total_in_kgce'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgce_per_prodution'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgco2e'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgco2e_per_prodution'))),
            'Subtotal',
            round_or_blank(report.get('reporting_total_production')),
        ])
    else:
        max_length = len(reporting_timestamps)
        for index in range(max_length):
            rows.append([
                reporting_timestamps[index],
                round_or_blank(reporting_values[index] if index < len(reporting_values) else None),
                round_or_blank(to_tons(values_in_kgce[index]) if index < len(values_in_kgce) else None),
                round_or_blank(to_tons(values_in_kgce_per_production[index]) if index < len(values_in_kgce_per_production) else None),
                round_or_blank(to_tons(values_in_kgco2e[index]) if index < len(values_in_kgco2e) else None),
                round_or_blank(to_tons(values_in_kgco2e_per_production[index]) if index < len(values_in_kgco2e_per_production) else None),
            ])

        rows.append([
            'Subtotal',
            round_or_blank(report.get('reporting_total_production')),
            round_or_blank(to_tons(reporting_period.get('total_in_kgce'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgce_per_prodution'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgco2e'))),
            round_or_blank(to_tons(reporting_period.get('total_in_kgco2e_per_prodution'))),
        ])

    return rows


def round_or_blank(value):
    if value is None:
        return ''
    return round2(value, 2)


def to_tons(value):
    if value is None:
        return None
    return value / 1000


def format_rate(value):
    if value is None:
        return '-'
    return str(round2(value * 100, 2)) + '%'


def format_unit(unit):
    if unit is None or len(str(unit)) == 0:
        return ''
    return ' (' + str(unit) + ')'


def is_base_period_timestamp_exists(base_period_data):
    timestamps = base_period_data.get('timestamps') if isinstance(base_period_data, dict) else None
    if timestamps is None or len(timestamps) == 0:
        return False

    for timestamp_list in timestamps:
        if timestamp_list is not None and len(timestamp_list) > 0:
            return True

    return False